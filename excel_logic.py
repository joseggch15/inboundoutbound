# excel_logic.py
# ============================================================
# Utilidades de Excel para PlanStaff con SSoT (SQLite),
# validación de estructura, importación validada,
# exportación/regeneración desde BD, comparación Excel↔BD,
# y utilidades de reporte.
#
# Esta versión corrige el error de Pylance:
#   "variant no está definido"
# garantizando que la variable 'variant' se inicializa y se
# propaga correctamente en validate_excel_structure y en
# cualquier flujo que la use (meta['variant']).
# ============================================================

from __future__ import annotations

import os
import io
from datetime import date, timedelta, datetime
from typing import List, Dict, Tuple, Optional, Set

import pandas as pd
import openpyxl
import xlsxwriter
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment

# ============================================================
# Helpers / Normalización
# ============================================================

def _is_blank_series(s: pd.Series) -> bool:
    """True si toda la serie es NaN o strings vacíos."""
    if s is None:
        return True
    if s.isna().all():
        return True
    s_str = s.astype(str).str.strip().str.lower()
    return (s_str.eq('') | s_str.eq('nan') | s_str.eq('none') | s_str.eq('null')).all()


def _prefix_for_file(plan_staff_file: str) -> str:
    """Prefijo de badge basado en el archivo."""
    base = os.path.basename(plan_staff_file).lower()
    if 'newmont' in base:
        return 'NM'
    return 'ID'


def _normalize_status(v: object) -> Tuple[Optional[str], Optional[str]]:
    """
    Normaliza celdas a ('ON'|'ON NS'|'OFF'|None, 'Day Shift'|'Night Shift'|None).
    - Valores numéricos o 'OK' se consideran ON (día).
    - 'ON' -> ON (día)
    - 'ON NS' o 'NIGHT' -> ON NS (noche)
    - 'OFF', 'Break', 'KO', 'Leave' -> OFF
    - Cualquier otro valor -> None (ignorar)
    Nota: Tipos personalizados (códigos) no se normalizan aquí (se tratan fuera).
    """
    if v is None:
        return None, None
    s = str(v).strip().upper()
    if s in ("", "NONE", "NAN", "NULL"):
        return None, None
    if s in ("OFF", "BREAK", "KO", "LEAVE"):
        return "OFF", None
    if s == "ON":
        return "ON", "Day Shift"
    if "ON NS" in s or "NIGHT" in s:
        return "ON NS", "Night Shift"
    # dígitos o 'OK'
    if s.isdigit() or s == "OK":
        return "ON", "Day Shift"
    return None, None


def _is_date_header(col) -> bool:
    """Detecta si el encabezado es una fecha (datetime o pandas.Timestamp)."""
    if isinstance(col, datetime):
        return True
    return getattr(col, '__class__', None).__name__ == 'Timestamp'


def _to_pydate(col) -> Optional[date]:
    """Convierte encabezado a date."""
    if isinstance(col, datetime):
        return col.date()
    try:
        return col.to_pydatetime().date()  # pandas.Timestamp
    except Exception:
        return None


def _fill_for_base_status(status: Optional[str]) -> Optional[PatternFill]:
    """Devuelve PatternFill para estados base ('ON', 'OFF', 'ON NS')."""
    if status is None:
        return None
    s = str(status).strip().upper()
    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # ON día
    red   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # OFF
    yel   = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # ON NS
    if s == "ON":
        return green
    if s == "OFF":
        return red
    if s == "ON NS":
        return yel
    return None

# =================================================================
# SHARED REPORTING HELPERS
# =================================================================

def _hhmmss(t: Optional[str]) -> Optional[str]:
    """Ensure time string is in HH:MM:SS format."""
    if t and len(t) == 5:
        return t + ":00"
    return t

def _get_transport_time_str(
    status: Optional[str],
    kind: str,
    comment: Optional[str],
    custom_map: Dict
) -> str:
    """
    Determines the transport time ('HH:MM:SS') based on shift status.
    'kind' is either 'IN' or 'OUT'.
    """
    su = (status or "").strip().upper()
    
    # Priority 1: Custom shift types from DB
    info = custom_map.get(su)
    if info:
        time_key = "in_time" if kind == "IN" else "out_time"
        t = info.get(time_key)
        if t:
            return _hhmmss(t) or "00:00:00"

    # Priority 2: Standard shifts
    if su == "ON":
        return "06:00:00" if kind == "IN" else "12:00:00"
    if su == "ON NS":
        return "12:00:00" if kind == "IN" else "06:00:00"

    # Priority 3: Fallback from cell comment (e.g., "08:00-17:00")
    if comment and "-" in str(comment):
        try:
            parts = str(comment).split("-", 1)
            t = parts[0].strip() if kind == "IN" else parts[1].strip()
            return _hhmmss(t) or "00:00:00"
        except (IndexError, AttributeError):
            pass

    # Final fallback
    return "06:00:00" if kind == "IN" else "12:00:00"


# ============================================================
# Lecturas auxiliares / previews
# ============================================================

def get_schedule_preview(plan_staff_file: str) -> pd.DataFrame:
    """
    Carga un DataFrame con columnas: ROLE, NAME, BADGE y columnas fecha.
    Aplica FR-07: limpia NaN/None en celdas de fechas para no mostrar 'nan'.
    Si el archivo no existe o falla, retorna df vacío.
    """
    if not os.path.exists(plan_staff_file):
        return pd.DataFrame()

    try:
        df = pd.read_excel(plan_staff_file, engine='openpyxl')
        # Detectar columnas base
        cols = list(df.columns)
        base_cols = []
        for c in ["ROLE", "NAME", "BADGE", "TEAM", "Discipline", "Company ID"]:
            if c in cols:
                base_cols.append(c)
        # Normalizar a ROLE/NAME/BADGE si vienen en formato Newmont
        if "Discipline" in base_cols and "Company ID" in base_cols:
            df["ROLE"] = df["Discipline"]
            if "BADGE" not in df.columns and "Company ID" in df.columns:
                df["BADGE"] = df["Company ID"]
        # Fechas
        date_cols = [c for c in cols if _is_date_header(c)]
        keep: List[str] = []
        # Mantener ROLE/NAME/BADGE si existen
        for key in ["ROLE", "NAME", "BADGE"]:
            if key in df.columns:
                keep.append(key)
        # si no había BADGE pero hay Company ID, también
        if "Company ID" in df.columns and "BADGE" not in keep:
            df["BADGE"] = df["Company ID"].astype(str)
            keep.append("BADGE")
        # armar df final
        keep += date_cols
        if not keep:
            return pd.DataFrame()

        df_out = df[keep].copy()

        # FR-07: limpiar NaN/None -> '' SOLO para visualización en la UI (no altera importación)
        for c in date_cols:
            if c in df_out.columns:
                df_out[c] = df_out[c].apply(
                    lambda v: "" if (v is None or str(v).strip().lower() in ("nan", "none", "null")) else v
                )

        return df_out
    except Exception:
        return pd.DataFrame()


def get_roles_from_excel(plan_staff_file: str) -> list:
    """Lista de roles únicos (si está disponible)."""
    if not os.path.exists(plan_staff_file):
        return [f"Role not available ({os.path.basename(plan_staff_file)} not found)"]
    try:
        wb = openpyxl.load_workbook(plan_staff_file, read_only=True, data_only=True)
        ws = wb.active
        header_map = {cell.value: cell.column for cell in ws[1]}
        role_header = "ROLE" if "ROLE" in header_map else ("Discipline" if "Discipline" in header_map else None)
        if not role_header:
            return ["ROLE/Discipline column not found"]
        col_idx = header_map[role_header]
        roles = {
            ws.cell(row=i, column=col_idx).value
            for i in range(2, ws.max_row + 1)
            if ws.cell(row=i, column=col_idx).value
        }
        return sorted(list(roles))
    except Exception:
        return ["Error reading Excel"]


def get_users_from_excel(plan_staff_file: str) -> list:
    """
    Extrae usuarios (name, role, badge) de la planilla.
    Soporta:
      - RGM: NAME, ROLE, BADGE
      - Newmont: Last Name, First Name, Discipline, Company ID
    Si no hay badge, genera uno estable (prefijo NM o ID + secuencia).
    """
    if not os.path.exists(plan_staff_file):
        return []
    try:
        df = pd.read_excel(plan_staff_file, engine='openpyxl')

        rgm_cols = ['NAME', 'ROLE', 'BADGE']
        newmont_cols = ['Last Name', 'First Name', 'Discipline', 'Company ID']

        users_df = None

        if all(col in df.columns for col in rgm_cols):
            users_df = df[rgm_cols].copy()
            # Badges faltantes
            if _is_blank_series(users_df['BADGE']):
                prefix = _prefix_for_file(plan_staff_file)
                users_df['BADGE'] = [f"{prefix}{i+1:05d}" for i in range(len(users_df))]
            else:
                prefix = _prefix_for_file(plan_staff_file)
                badge_series = users_df['BADGE'].astype(str)
                is_missing = users_df['BADGE'].isna() | badge_series.str.strip().eq('') | badge_series.str.lower().isin(['nan', 'none', 'null'])
                seq = (f"{prefix}{i+1:05d}" for i in range(is_missing.sum()))
                users_df.loc[is_missing, 'BADGE'] = [next(seq) for _ in range(is_missing.sum())]

        elif all(col in df.columns for col in newmont_cols):
            df_copy = df[newmont_cols].copy()
            df_copy['NAME'] = df_copy['Last Name'].astype(str).str.strip() + ', ' + df_copy['First Name'].astype(str).str.strip()
            df_copy.rename(columns={'Discipline': 'ROLE', 'Company ID': 'BADGE'}, inplace=True)
            users_df = df_copy[['NAME', 'ROLE', 'BADGE']]
            if _is_blank_series(users_df['BADGE']):
                prefix = _prefix_for_file(plan_staff_file)
                users_df['BADGE'] = [f"{prefix}{i+1:05d}" for i in range(len(users_df))]
        else:
            # Fallback si vienen NAME/ROLE solamente
            if all(col in df.columns for col in ['NAME', 'ROLE']):
                prefix = _prefix_for_file(plan_staff_file)
                users_df = df[['NAME', 'ROLE']].copy()
                users_df['BADGE'] = [f"{prefix}{i+1:05d}" for i in range(len(users_df))]
            else:
                return []

        users_df['NAME'] = users_df['NAME'].astype(str).str.strip()
        users_df['ROLE'] = users_df['ROLE'].astype(str).str.strip()
        users_df['BADGE'] = users_df['BADGE'].astype(str).str.strip()

        users_df = users_df[(users_df['NAME'] != '') & (users_df['BADGE'] != '')]
        users_df.drop_duplicates(subset=['BADGE'], keep='first', inplace=True)
        users_df.rename(columns={'NAME': 'name', 'ROLE': 'role', 'BADGE': 'badge'}, inplace=True)
        return users_df.to_dict('records')
    except Exception:
        return []


# ============================================================
# Escritura / actualización del plan staff (Excel)
# ============================================================

def update_plan_staff_excel(
    plan_staff_file: str,
    username: str,
    role: str,
    badge: str,
    schedule_status: Optional[str],
    shift_type: Optional[str],
    schedule_start: date,
    schedule_end: date,
    source: str,
    in_time: Optional[str] = None,
    out_time: Optional[str] = None
) -> Tuple[bool, str]:
    """
    Actualiza (o crea si no existe) la fila del empleado en el Excel:
    - Busca por BADGE y, si no, por NAME.
    - Escribe:
        * Estados base -> 'ON' / 'ON NS' / 'OFF' con colores legacy.
        * Tipos personalizados -> código (p.ej. 'SOP') y color del tipo.
      Además añade un comentario con 'IN-OUT' (HH:MM-HH:MM) si viene in_time/out_time.
    - Si schedule_status es None, limpia el rango.
    """
    try:
        # Abrir o crear
        if os.path.exists(plan_staff_file):
            wb = openpyxl.load_workbook(plan_staff_file)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Operations_best_opt"
            headers = ["TEAM", "ROLE", "NAME", "BADGE"]
            for col_idx, h in enumerate(headers, start=1):
                ws.cell(row=1, column=col_idx, value=h)

        # Colores base
        green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # ON día
        red   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # OFF
        yel   = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # ON NS noche

        # Mapa de tipos personalizados (para colores)
        try:
            from database_logic import get_shift_type_map  # import diferido para evitar ciclos
            custom_map = get_shift_type_map(source)
        except Exception:
            custom_map = {}

        def _fill_for_status(status: Optional[str]) -> Optional[PatternFill]:
            if status is None:
                return None
            s = str(status).strip().upper()
            if s == "ON":
                return green
            if s == "ON NS":
                return yel
            if s == "OFF":
                return red
            # código personalizado
            info = custom_map.get(s)
            if info and info.get('color_hex'):
                hex6 = info['color_hex'].lstrip('#').upper()
                return PatternFill(start_color=hex6, end_color=hex6, fill_type="solid")
            return None

        # Mapas de cabecera
        header_map = {cell.value: cell.column for cell in ws[1] if isinstance(cell.value, str)}
        date_map: Dict[date, int] = {}
        for cell in ws[1]:
            v = cell.value
            if isinstance(v, datetime):
                date_map[v.date()] = cell.column

        # localizar fila por BADGE o NAME
        row_idx = None
        if "BADGE" in header_map:
            for i in range(2, ws.max_row + 1):
                v = ws.cell(row=i, column=header_map["BADGE"]).value
                if v and str(v) == str(badge):
                    row_idx = i
                    break
        if not row_idx and "NAME" in header_map:
            for i in range(2, ws.max_row + 1):
                v = ws.cell(row=i, column=header_map["NAME"]).value
                if v and str(v).strip().lower() == str(username).strip().lower():
                    row_idx = i
                    break
        if not row_idx:
            row_idx = ws.max_row + 1

        # Datos fijos
        if "NAME" in header_map:
            ws.cell(row=row_idx, column=header_map["NAME"], value=username)
        if "ROLE" in header_map:
            ws.cell(row=row_idx, column=header_map["ROLE"], value=role)
        if "BADGE" in header_map:
            ws.cell(row=row_idx, column=header_map["BADGE"], value=badge)

        # texto/estilo por estado
        text = None
        fill = None
        # status can now be a custom code like 'SOP'
        if schedule_status:
            text = str(schedule_status).strip().upper()
            fill = _fill_for_status(text)
        else: # clearing range
            text = None
            fill = None

        # Escribir/limpiar rango
        d = schedule_start
        while d <= schedule_end:
            # Crear columna de fecha si no existe en el template
            if d not in date_map:
                new_col = ws.max_column + 1
                ws.cell(row=1, column=new_col, value=datetime(d.year, d.month, d.day))
                date_map[d] = new_col
            col_idx = date_map[d]
            cell = ws.cell(row=row_idx, column=col_idx)
            if text is None:
                cell.value = None
                cell.comment = None
                cell.fill = PatternFill(fill_type=None)
            else:
                cell.value = text
                cell.fill = fill if fill else PatternFill(fill_type=None)
                # Comentario con horarios para personalizados
                if text not in ("ON", "ON NS", "OFF") and in_time and out_time:
                    cell.comment = Comment(f"{in_time}-{out_time}", "ShiftType")
                else:
                    cell.comment = None
            d += timedelta(days=1)

        wb.save(plan_staff_file)
        return True, f"Plan staff updated for {username}."
    except Exception as e:
        return False, f"Error updating plan staff: {e}"


# ============================================================
# FR-01: Detección de conflictos (sobrescritura)
# ============================================================

def find_conflicts(plan_staff_file: str, username: str, badge: str,
                   schedule_start: date, schedule_end: date) -> List[Dict]:
    """
    Devuelve [{'date': date, 'existing': 'ON/ON NS/OFF/...'}] si hay valores ya escritos en el rango.
    Busca fila por BADGE y luego por NAME, igual que update_plan_staff_excel.
    """
    try:
        if not os.path.exists(plan_staff_file):
            return []
        wb = openpyxl.load_workbook(plan_staff_file, data_only=True)
        ws = wb.active

        header_map = {cell.value: cell.column for cell in ws[1] if isinstance(cell.value, str)}
        date_map = {cell.value.date(): cell.column for cell in ws[1] if isinstance(cell.value, datetime)}

        # localizar fila por BADGE y luego por NAME
        row_idx = None
        if "BADGE" in header_map:
            for i in range(2, ws.max_row + 1):
                v = ws.cell(row=i, column=header_map["BADGE"]).value
                if v and str(v) == str(badge):
                    row_idx = i
                    break
        if not row_idx and "NAME" in header_map:
            for i in range(2, ws.max_row + 1):
                v = ws.cell(row=i, column=header_map["NAME"]).value
                if v and str(v).strip().lower() == str(username).strip().lower():
                    row_idx = i
                    break
        if not row_idx:
            return []

        conflicts: List[Dict] = []
        d = schedule_start
        while d <= schedule_end:
            if d in date_map:
                col = date_map[d]
                val = ws.cell(row=row_idx, column=col).value
                if val not in (None, '', ' '):
                    conflicts.append({"date": d, "existing": str(val)})
            d += timedelta(days=1)
        return conflicts
    except Exception:
        return []


# ============================================================
# FR-02: Importar Excel -> DB (usuarios + schedules) con validación
# ============================================================

def import_excel_to_db(plan_staff_file: str, source: str) -> Tuple[int, int, int]:
    """
    Procesa .xlsx y almacena en la BD:
      - Usuarios (name, role, badge)
      - Schedules día-a-día (ON/ON NS/OFF)
    Devuelve: (nuevos_usuarios, usuarios_omitidos, upserts_schedule)

    Si el archivo NO es válido (estructura), levanta ValueError con detalle.
    """
    # Validación previa estricta
    ok, errors, _meta = validate_excel_structure(plan_staff_file)
    if not ok:
        raise ValueError("Invalid Plan Staff structure:\n" + "\n".join(f"- {e}" for e in errors))

    from database_logic import add_users_bulk, get_all_users, upsert_schedule_day  # import diferido

    users_in_file = get_users_from_excel(plan_staff_file)
    if not users_in_file:
        return (0, 0, 0)

    # Insertar usuarios (evitando duplicados por badge)
    before = get_all_users(source)
    before_badges = {u['badge'] for u in before}
    inserted = add_users_bulk(users_in_file, source)
    after = get_all_users(source)
    after_badges = {u['badge'] for u in after}
    skipped = len(before_badges & after_badges)  # aproximado para el mensaje

    # ---- Validación de tipos de turno personalizados (códigos) ----
    # Recorremos todas las columnas de fechas y recopilamos valores que no sean
    # estados base ('ON', 'ON NS', 'OFF') ni equivalentes a ON por números/OK/DAY/NIGHT.
    # Cualquier otro valor se considera un "código" de turno que debe existir en shift_types.
    try:
        from database_logic import get_shift_type_map
        _custom_map = {k.strip().upper(): v for k,v in get_shift_type_map(source).items()}
    except Exception:
        _custom_map = {}
    try:
        df_codes = pd.read_excel(plan_staff_file, engine='openpyxl')
        date_cols_all = [c for c in df_codes.columns if _is_date_header(c)]
        unknown_codes = set()
        for dcol in date_cols_all:
            col_series = df_codes[dcol]
            for v in col_series:
                if v is None:
                    continue
                s = str(v).strip()
                if not s:
                    continue
                su = s.upper()
                # equivalencias/estados base
                if su in ("ON", "ON NS", "OFF", "BREAK", "KO", "LEAVE"):
                    continue
                if su.isdigit() or su == "OK" or ("DAY" in su) or ("NIGHT" in su):
                    # se tratará como ON (día/noche) -> no requiere tipo personalizado
                    continue
                # Si llega aquí lo tratamos como código personalizado
                if su not in _custom_map:
                    unknown_codes.add(su)
        if unknown_codes:
            # abortar importación (la UI capturará este ValueError y lo mostrará en un QMessageBox)
            raise ValueError(
                "Se detectaron turnos/códigos no registrados en 'shift_types':\n  - " +
                "\n  - ".join(sorted(unknown_codes)) +
                "\n\nRegístrelos primero (nombre, código y horarios IN/OUT) en 'Shift Types' para continuar."
            )
    except ValueError:
        # re-lanzar para que la capa UI lo muestre
        raise
    except Exception:
        # fallas al leer códigos no deben romper la importación; continuamos
        pass

    # Importar schedules (solo estados base reconocidos)
    upserts = 0
    try:
        df = pd.read_excel(plan_staff_file, engine='openpyxl')
        # detectar identificadores
        role_field = 'ROLE' if 'ROLE' in df.columns else ('Discipline' if 'Discipline' in df.columns else None)
        badge_field = 'BADGE' if 'BADGE' in df.columns else ('Company ID' if 'Company ID' in df.columns else None)

        if not badge_field:
            return inserted, skipped, 0

        date_cols = [c for c in df.columns if _is_date_header(c)]
        for _, row in df.iterrows():
            badge = str(row[badge_field]).strip()
            if not badge:
                continue
            for dcol in date_cols:
                d_py = _to_pydate(dcol)
                if not d_py:
                    continue
                status, shift = _normalize_status(row[dcol])
                if status:
                    upsert_schedule_day(badge, d_py, status, shift, source)
                    upserts += 1
    except Exception:
        pass

    return (inserted, skipped, upserts)


# ============================================================
# FR-03: Exportar Excel desde la BD (manteniendo plantilla)
# ============================================================

def export_plan_from_db(
    template_path: str,
    users: List[Dict],
    schedules: List[Dict],
    output_path: str,
    source: str
) -> Tuple[bool, str]:
    """
    users: [{'name','role','badge'}]
    schedules: [{'badge','date':'YYYY-MM-DD','status','shift_type','in_time','out_time'}]
    Soporta plantillas RGM (NAME/ROLE/BADGE) y Newmont (Last/First/Discipline/Company ID).
    """
    try:
        if not os.path.exists(template_path):
            return False, f"Template '{os.path.basename(template_path)}' not found."

        try:
            from database_logic import get_shift_type_map  # import diferido
            custom_map = get_shift_type_map(source)
        except Exception:
            custom_map = {}

        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        header_map = {cell.value: cell.column for cell in ws[1] if isinstance(cell.value, str)}

        # Detectar variante de plantilla
        variant = None
        if all(h in header_map for h in ("NAME", "ROLE", "BADGE")):
            variant = "RGM"
        elif all(h in header_map for h in ("Last Name", "First Name", "Discipline", "Company ID")):
            variant = "Newmont"
        else:
            return False, "Unsupported template: expected RGM (NAME/ROLE/BADGE) or Newmont (Last/First/Discipline/Company ID)."

        # Mapas de fechas existentes
        date_map: Dict[date, int] = {}
        for c in ws[1]:
            if isinstance(c.value, datetime):
                date_map[c.value.date()] = c.column

        # Colores base
        green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        yel   = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

        def _fill_for(status: Optional[str]) -> Optional[PatternFill]:
            if status is None:
                return None
            s = str(status).strip().upper()
            if s == "OFF":
                return red
            if s == "ON NS":
                return yel
            if s == "ON":
                return green
            info = custom_map.get(s)
            if info and info.get('color_hex'):
                hex6 = info['color_hex'].lstrip('#').upper()
                return PatternFill(start_color=hex6, end_color=hex6, fill_type="solid")
            return None

        # Schedules -> mapa por badge y fecha
        sched_by_badge: Dict[str, Dict[str, Dict[str, Optional[str]]]] = {}
        for s in schedules:
            b = str(s.get('badge', '')).strip()
            d = str(s.get('date', '')).strip()  # 'YYYY-MM-DD'
            if not b or not d:
                continue
            sched_by_badge.setdefault(b, {})[d] = {
                'status': s.get('status'),
                'shift_type': s.get('shift_type'),
                'in_time': s.get('in_time'),
                'out_time': s.get('out_time')
            }

        # Helper para nombre en Newmont
        def split_name(full: str) -> Tuple[str, str]:
            if not full:
                return "", ""
            s = str(full).strip()
            if "," in s:
                last, first = s.split(",", 1)
                return last.strip(), first.strip()
            parts = s.split()
            if len(parts) >= 2:
                return parts[-1], " ".join(parts[:-1])
            return s, ""

        # Escribir/actualizar usuarios
        # Índice rápido por badge ya existente
        badge_col = header_map["BADGE"] if variant == "RGM" else header_map["Company ID"]
        existing_rows_by_badge: Dict[str, int] = {}
        for i in range(2, ws.max_row + 1):
            v = ws.cell(row=i, column=badge_col).value
            if v:
                existing_rows_by_badge[str(v).strip()] = i

        for u in users:
            name = u.get('name', '').strip()
            role = u.get('role', '').strip()
            badge = str(u.get('badge', '')).strip()
            if not badge:
                continue

            row_idx = existing_rows_by_badge.get(badge)
            if not row_idx:
                row_idx = ws.max_row + 1
                existing_rows_by_badge[badge] = row_idx

            if variant == "RGM":
                ws.cell(row=row_idx, column=header_map["NAME"], value=name)
                ws.cell(row=row_idx, column=header_map["ROLE"], value=role)
                ws.cell(row=row_idx, column=header_map["BADGE"], value=badge)
            else:
                last, first = split_name(name)
                ws.cell(row=row_idx, column=header_map["Last Name"], value=last)
                ws.cell(row=row_idx, column=header_map["First Name"], value=first)
                ws.cell(row=row_idx, column=header_map["Discipline"], value=role)
                ws.cell(row=row_idx, column=header_map["Company ID"], value=badge)

            # Rellenar días
            # Fechas ordenadas por las que ya existen en plantilla
            dates_sorted = sorted(date_map.keys())
            for d in dates_sorted:
                col_idx = date_map[d]
                cell = ws.cell(row=row_idx, column=col_idx)
                info = sched_by_badge.get(badge, {}).get(d.isoformat(), None)
                if info:
                    st = (info.get('status') or '').strip().upper() if info.get('status') else None
                    cell.value = st
                    cell.fill = _fill_for(st) or PatternFill(fill_type=None)
                    # comentario para personalizados si tenemos HH:MM
                    if st and st not in ("ON", "ON NS", "OFF"):
                        it = (info.get('in_time') or '').strip()
                        ot = (info.get('out_time') or '').strip()
                        if it and ot:
                            cell.comment = Comment(f"{it}-{ot}", "ShiftType")
                        else:
                            cell.comment = None
                else:
                    cell.value = None
                    cell.fill = PatternFill(fill_type=None)
                    cell.comment = None

        # Expandir columnas si hay fechas en BD que no existían en plantilla
        # (opcional; se puede agregar al final)
        all_sched_dates: Set[date] = set()
        for b, per_day in sched_by_badge.items():
            for d_str in per_day.keys():
                try:
                    y, m, dy = d_str.split("-")
                    all_sched_dates.add(date(int(y), int(m), int(dy)))
                except Exception:
                    pass
        missing_dates = sorted([d for d in all_sched_dates if d not in date_map])
        for d in missing_dates:
            new_col = ws.max_column + 1
            ws.cell(row=1, column=new_col, value=datetime(d.year, d.month, d.day))
            date_map[d] = new_col
            # escribir valores para cada usuario
            for u in users:
                badge = str(u.get('badge', '')).strip()
                if not badge:
                    continue
                row_idx = existing_rows_by_badge.get(badge)
                if not row_idx:
                    continue
                info = sched_by_badge.get(badge, {}).get(d.isoformat(), None)
                cell = ws.cell(row=row_idx, column=new_col)
                if info:
                    st = (info.get('status') or '').strip().upper() if info.get('status') else None
                    cell.value = st
                    cell.fill = _fill_for(st) or PatternFill(fill_type=None)
                    if st and st not in ("ON", "ON NS", "OFF"):
                        it = info.get('in_time')
                        ot = info.get('out_time')
                        if it and ot:
                            cell.comment = Comment(f"{it}-{ot}", "ShiftType")
                        else:
                            cell.comment = None
                else:
                    cell.value = None
                    cell.fill = PatternFill(fill_type=None)
                    cell.comment = None

        wb.save(output_path)
        return True, f"Plan successfully exported to '{os.path.basename(output_path)}'."
    except Exception as e:
        return False, f"Export error: {e}"


# ============================================================
# Reporte de transporte (IN/OUT) — MODIFIED to use xlsxwriter
# ============================================================

def generate_transport_report(
    plan_staff_file: str,
    start_date: date,
    end_date: date,
    settings: Dict
) -> Tuple[bytes, str]:
    """
    Generates the transport Excel report using xlsxwriter to apply custom formatting.
    This now defaults to the new layout for Newmont as per the image.
    """
    # ---- 1) Inferir source por nombre de archivo ----
    fname = os.path.basename(plan_staff_file).lower()
    source = "Newmont" if "newmont" in fname else "RGM"
    
    if source == "RGM":
        return generate_rgm_transport_report(plan_staff_file, start_date, end_date, settings)

    # ---- 2) Cargar mapping de tipos personalizados desde BD ----
    try:
        from database_logic import get_shift_type_map, get_user_location_for_date
        custom_map: Dict[str, Dict] = {k.strip().upper(): v for k, v in get_shift_type_map(source).items()}
    except Exception:
        custom_map = {}
        def get_user_location_for_date(b, d): return (None, None)

    # ---- 3) Abrir plan staff y extraer datos ----
    try:
        wb_src = openpyxl.load_workbook(plan_staff_file, data_only=True)
        ws_src = wb_src.active
    except Exception as e:
        return b"", f"Could not read Plan Staff file: {e}"

    # ---- 4) Resolver esquema de columnas (RGM vs Newmont) ----
    header_map: Dict[str, int] = {c.value: c.column for c in ws_src[1] if isinstance(c.value, str)}
    date_cols: Dict[int, date] = {c.column: c.value.date() for c in ws_src[1] if isinstance(c.value, datetime)}

    is_rgm = all(h in header_map for h in ("NAME", "ROLE", "BADGE"))
    is_new = all(h in header_map for h in ("Last Name", "First Name", "Discipline", "Company ID"))
    if not (is_rgm or is_new):
        return b"", "Unsupported Plan Staff format."

    # ---- 5) Helpers and Data Extraction ----
    OFF_LIKE = {"OFF", "BREAK", "KO", "LEAVE"}
    def _norm_status(v):
        if v is None: return None
        s = str(v).strip()
        if not s: return None
        su = s.upper()
        if su in OFF_LIKE: return "OFF"
        if su in ("ON", "ON NS"): return su
        if su.isdigit() or su == "OK" or "DAY" in su: return "ON"
        return su
    def _is_working(s): return bool(s and s not in ("OFF", "BREAK", "KO", "LEAVE"))

    # ---- 6) Data processing: Collect IN/OUT rows ----
    in_rows_data, out_rows_data = [], []
    company_default = "PLGims"
    dates_sorted = sorted(date_cols.values())

    if is_rgm:
        name_col, role_col, badge_col = header_map["NAME"], header_map["ROLE"], header_map["BADGE"]
        def get_name(r):
            nm = str(ws_src.cell(row=r, column=name_col).value or "").strip()
            if "," in nm: last, first = nm.split(",", 1); return last.strip(), first.strip()
            parts = nm.split(); return (parts[-1], " ".join(parts[:-1])) if len(parts) >= 2 else (nm, "")
    else:
        ln_col, fn_col, role_col, badge_col = header_map["Last Name"], header_map["First Name"], header_map["Discipline"], header_map["Company ID"]
        def get_name(r):
            ln = str(ws_src.cell(row=r, column=ln_col).value or "").strip()
            fn = str(ws_src.cell(row=r, column=fn_col).value or "").strip()
            return ln, fn

    for r_idx in range(2, ws_src.max_row + 1):
        badge = str(ws_src.cell(row=r_idx, column=badge_col).value or "").strip()
        if not badge: continue
        role, (last, first) = str(ws_src.cell(row=r_idx, column=role_col).value or "").strip(), get_name(r_idx)
        
        per_day: Dict[date, Tuple[Optional[str], Optional[str]]] = {}
        for c, d in date_cols.items():
            cell = ws_src.cell(row=r_idx, column=c)
            per_day[d] = (_norm_status(cell.value), (cell.comment.text if cell.comment else None))

        if not per_day: continue
        added_in, added_out = set(), set()
        next_in, next_out = None, None

        for i, d in enumerate(dates_sorted):
            st_d, cmt_d = per_day.get(d, (None, None))
            if not _is_working(st_d): continue
            prev_d = dates_sorted[i-1] if i > 0 else None
            next_d = dates_sorted[i+1] if i < len(dates_sorted)-1 else None
            st_prev, _ = per_day.get(prev_d, (None,None)) if prev_d else (None,None)
            st_next, _ = per_day.get(next_d, (None,None)) if next_d else (None,None)

            time_in_str = _get_transport_time_str(st_d, "IN", cmt_d, custom_map)
            time_out_str = _get_transport_time_str(st_d, "OUT", cmt_d, custom_map)

            if st_prev != st_d: # Entry event
                if start_date <= d <= end_date and d not in added_in:
                    pu, _ = get_user_location_for_date(badge, d)
                    in_rows_data.append([last, first, badge, company_default, role, pu or "", d, time_in_str])
                    added_in.add(d)
                elif d > end_date and next_in is None:
                    next_in = (d, st_d, cmt_d)
            if st_next != st_d: # Exit event
                if start_date <= d <= end_date and d not in added_out:
                    _, do = get_user_location_for_date(badge, d)
                    out_rows_data.append([last, first, badge, company_default, role, do or "", d, time_out_str])
                    added_out.add(d)
                elif d > end_date and next_out is None:
                    next_out = (d, st_d, cmt_d)

        if next_in and next_in[0] not in added_in:
            d, st, cmt = next_in; pu, _ = get_user_location_for_date(badge, d)
            time_in_str = _get_transport_time_str(st, "IN", cmt, custom_map)
            in_rows_data.append([last, first, badge, company_default, role, pu or "", d, time_in_str])
        if next_out and next_out[0] not in added_out:
            d, st, cmt = next_out; _, do = get_user_location_for_date(badge, d)
            time_out_str = _get_transport_time_str(st, "OUT", cmt, custom_map)
            out_rows_data.append([last, first, badge, company_default, role, do or "", d, time_out_str])

    # ---- 7) Write to xlsxwriter workbook ----
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    ws = workbook.add_worksheet("travel list")

    # --- Formats ---
    base_font_settings = {'font_name': settings.get('font_name', 'Calibri'), 'valign': 'vcenter'}
    
    f_title = workbook.add_format({'bold': True, 'font_name': base_font_settings['font_name'], 'underline': True, 'font_size': 11})
    
    f_header = workbook.add_format({
        'bold': True, 
        'font_name': base_font_settings['font_name'], 
        'font_color': settings.get('header_font_color'), 
        'bg_color': settings.get('header_bg_color'), 
        'align': 'center', 'valign': 'vcenter', 'border': 1
    })
    
    f_default = workbook.add_format({**base_font_settings, 'border': 1})
    f_date = workbook.add_format({**base_font_settings, 'border': 1, 'num_format': settings.get('date_format', 'dd-mmm-yy'), 'font_color': 'red'})
    f_time = workbook.add_format({**base_font_settings, 'border': 1, 'num_format': 'hh:mm', 'font_color': 'red'})

    # --- Write Content ---
    ws.write('A1', "MERIAN TRANSPORTATION REQUEST", f_title)
    ws.write('A3', "IN", workbook.add_format({'bold': True}))
    ws.write('B3', "TRAVEL TO SITE", workbook.add_format({'bold': True}))
    ws.write('K3', "OUT", workbook.add_format({'bold': True}))
    ws.write('L3', "TRAVEL FROM SITE", workbook.add_format({'bold': True}))

    headers_in = ["#", "NAME", "FIRST NAME", "GID", "COMPANY", "DEPT", "FROM", "DATE", "TIME"]
    headers_out = ["#", "NAME", "FIRST NAME", "GID", "COMPANY", "DEPT", "TO", "DATE", "TIME"]
    for i, h in enumerate(headers_in): ws.write(3, i, h, f_header)
    for i, h in enumerate(headers_out): ws.write(3, i + 10, h, f_header)

    row_start_index = 4
    for i, row_data in enumerate(in_rows_data):
        ws.write(i + row_start_index, 0, i + 1, f_default)
        for j, cell_data in enumerate(row_data):
            col = j + 1
            if headers_in[col] == "DATE":
                ws.write_datetime(i + row_start_index, col, cell_data, f_date)
            elif headers_in[col] == "TIME":
                time_obj = datetime.strptime(cell_data, '%H:%M:%S')
                ws.write_datetime(i + row_start_index, col, time_obj, f_time)
            else:
                ws.write(i + row_start_index, col, cell_data, f_default)
                
    for i, row_data in enumerate(out_rows_data):
        ws.write(i + row_start_index, 10, i + 1, f_default)
        for j, cell_data in enumerate(row_data):
            col = j + 11
            if headers_out[col - 10] == "DATE":
                ws.write_datetime(i + row_start_index, col, cell_data, f_date)
            elif headers_out[col - 10] == "TIME":
                time_obj = datetime.strptime(cell_data, '%H:%M:%S')
                ws.write_datetime(i + row_start_index, col, time_obj, f_time)
            else:
                 ws.write(i + row_start_index, col, cell_data, f_default)
    
    ws.autofit()
    workbook.close()
    output.seek(0)
    return output.read(), "Transportation report generated."


# ============================================================
# NEW RGM-specific Report
# ============================================================
def generate_rgm_transport_report(plan_staff_file: str, start_date: date, end_date: date, settings: Dict) -> Tuple[bytes, str]:
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    worksheet = workbook.add_worksheet("Sheet1")

    # Formats
    title_format = workbook.add_format({'bold': True, 'font_color': 'red', 'align': 'center', 'valign': 'vcenter', 'font_size': 18})
    header_format = workbook.add_format({'bold': True, 'bg_color': '#00B0F0', 'font_color': 'white', 'align': 'center', 'valign': 'vcenter', 'border': 1})
    yellow_header_format = workbook.add_format({'bold': True, 'bg_color': 'yellow', 'align': 'center', 'valign': 'vcenter', 'border': 1})
    data_format = workbook.add_format({'align': 'left', 'valign': 'vcenter', 'border': 1})
    time_format = workbook.add_format({'align': 'left', 'valign': 'vcenter', 'border': 1, 'num_format': 'h:mm AM/PM'})
    date_format = workbook.add_format({'align': 'left', 'valign': 'vcenter', 'border': 1, 'num_format': settings.get('date_format', 'dd/mm/yyyy')})
    
    # Set column widths
    worksheet.set_column('A:A', 5)
    worksheet.set_column('B:B', 25) # NAME
    worksheet.set_column('C:D', 15) # DEPARTMENT, BADGE
    worksheet.set_column('E:F', 20) # POSITION, CREW
    worksheet.set_column('G:H', 25) # PICKUP, INBOUND DATE
    worksheet.set_column('I:K', 15) # METHOD, LOCATION, DEPT TIME

    worksheet.set_column('M:M', 25) # NAME
    worksheet.set_column('N:P', 15) # DEPARTMENT, BADGE, POSITION
    worksheet.set_column('Q:R', 20) # CREW, OUTBOUND DATE
    worksheet.set_column('S:T', 25) # METHOD, LOCATION
    worksheet.set_column('U:U', 15) # DEPT TIME

    # Inbound Section
    worksheet.merge_range('A1:K2', 'INBOUND', title_format)
    inbound_headers = ["NR", "NAME (Last, First Name)", "DEPARTMENT", "BADGE #", "POSITION / TITLE", "CREW A/B/C", "PICK UP LOCATION", "IN BOUND DATE", "Method Of Transport", "Location", "DEPT TIME"]
    for col, header in enumerate(inbound_headers):
        fmt = yellow_header_format if header == "PICK UP LOCATION" else header_format
        worksheet.write(2, col, header, fmt)

    # Outbound Section
    worksheet.merge_range('M1:U2', 'OUTBOUND', title_format)
    outbound_headers = ["NAME (Last, First Name)", "DEPARTMENT", "BADGE #", "POSITION / TITLE", "CREW A/B/C", "ROSEBEL SITE OUT BOUND DATE", "Method Of Transport", "Location", "DEPT TIME"]
    for col, header in enumerate(outbound_headers):
        worksheet.write(2, col + 12, header, header_format)

    # --- Data Extraction Logic (similar to the original function) ---
    try:
        from database_logic import get_shift_type_map, get_user_location_for_date
        custom_map: Dict[str, Dict] = {k.strip().upper(): v for k, v in get_shift_type_map("RGM").items()}
    except Exception:
        custom_map = {}
        def get_user_location_for_date(b, d): return (None, None)

    try:
        wb_src = openpyxl.load_workbook(plan_staff_file, data_only=True)
        ws_src = wb_src.active
    except Exception as e:
        workbook.close()
        return b"", f"Could not read Plan Staff file: {e}"

    header_map: Dict[str, int] = {c.value: c.column for c in ws_src[1] if isinstance(c.value, str)}
    date_cols: Dict[int, date] = {c.column: c.value.date() for c in ws_src[1] if isinstance(c.value, datetime)}
    
    OFF_LIKE = {"OFF", "BREAK", "KO", "LEAVE"}
    def _norm_status(v):
        if v is None: return None
        s = str(v).strip();
        if not s: return None
        su = s.upper()
        if su in OFF_LIKE: return "OFF"
        if su.isdigit() or su in ("OK", "ON") or "DAY" in su: return "ON"
        if "ON NS" in su or "NIGHT" in su: return "ON NS"
        return su
    def _is_working(s): return bool(s and s != "OFF")

    def _get_crew_from_name(name: str):
        # Placeholder logic
        if 'day' in name.lower(): return "A 14/7 DAY"
        if 'night' in name.lower(): return "B 7/7/7 DAY/NIGHT"
        return "C 14/7 DAY"

    in_row, out_row = 3, 3
    dates_sorted = sorted(date_cols.values())

    for r_idx in range(2, ws_src.max_row + 1):
        badge = str(ws_src.cell(row=r_idx, column=header_map["BADGE"]).value or "").strip()
        if not badge: continue

        name = str(ws_src.cell(row=r_idx, column=header_map["NAME"]).value or "")
        department = str(ws_src.cell(row=r_idx, column=header_map["ROLE"]).value or "")
        position = "Technician" # Placeholder
        
        per_day: Dict[date, Tuple[Optional[str], Optional[str]]] = {}
        for c, d in date_cols.items():
            cell = ws_src.cell(row=r_idx, column=c)
            per_day[d] = (_norm_status(cell.value), (cell.comment.text if cell.comment else None))

        for i, d in enumerate(dates_sorted):
            if not (start_date <= d <= end_date): continue
            
            st_d, cmt_d = per_day.get(d, (None, None))
            if not _is_working(st_d): continue

            prev_d = dates_sorted[i-1] if i > 0 else None
            st_prev, _ = per_day.get(prev_d, (None, None))
            
            # INBOUND event
            if st_prev != st_d:
                pu, _ = get_user_location_for_date(badge, d)
                crew = _get_crew_from_name(st_d if st_d else "")
                time_str = _get_transport_time_str(st_d, "IN", cmt_d, custom_map)
                dept_time = datetime.strptime(time_str, "%H:%M:%S")
                
                in_data = [in_row - 2, name, department, badge, position, crew, pu or "N/A", d, "RGM TRANSPORT", "PARAMARIBO", dept_time]
                for col, val in enumerate(in_data):
                    header_name = inbound_headers[col]
                    if header_name == "IN BOUND DATE":
                        worksheet.write_datetime(in_row, col, val, date_format)
                    elif header_name == "DEPT TIME":
                        worksheet.write_datetime(in_row, col, val, time_format)
                    else:
                        worksheet.write(in_row, col, val, data_format)
                in_row += 1

            # OUTBOUND event
            next_d = dates_sorted[i+1] if i < len(dates_sorted) - 1 else None
            st_next, _ = per_day.get(next_d, (None, None))

            if st_next != st_d:
                _, do = get_user_location_for_date(badge, d)
                crew = _get_crew_from_name(st_d if st_d else "")
                time_str = _get_transport_time_str(st_d, "OUT", cmt_d, custom_map)
                dept_time = datetime.strptime(time_str, "%H:%M:%S")

                out_data = [name, department, badge, position, crew, d, "RGM TRANSPORT", do or "PARAMARIBO", dept_time]
                for col, val in enumerate(out_data):
                    header_name = outbound_headers[col]
                    if header_name == "ROSEBEL SITE OUT BOUND DATE":
                        worksheet.write_datetime(out_row, col + 12, val, date_format)
                    elif header_name == "DEPT TIME":
                        worksheet.write_datetime(out_row, col + 12, val, time_format)
                    else:
                         worksheet.write(out_row, col + 12, val, data_format)
                out_row += 1


    workbook.close()
    output.seek(0)
    return output.read(), "RGM Transportation report generated."


# ============================================================
# Utilidad: Propagar cambios de código/color a Excel (inmediato)
# ============================================================

def apply_shift_type_update_to_excel(plan_staff_file: str, source: str, old_code: str, new_code: str, color_hex: str) -> Tuple[bool, str]:
    """
    Reemplaza en TODO el archivo Excel el código viejo por el nuevo y aplica el color indicado.
    No altera comentarios ni otros contenidos.
    """
    try:
        if not os.path.exists(plan_staff_file):
            return False, f"Plan staff '{os.path.basename(plan_staff_file)}' not found."

        wb = openpyxl.load_workbook(plan_staff_file)
        ws = wb.active

        hex6 = color_hex.lstrip('#').upper()
        fill = PatternFill(start_color=hex6, end_color=hex6, fill_type="solid")

        # Hallar columnas de fecha para no tocar cabeceras no relacionadas
        date_cols = set()
        for c in ws[1]:
            if isinstance(c.value, datetime):
                date_cols.add(c.column)

        # Buscar celdas con old_code y reemplazar (solo en columnas de fecha)
        for r in range(2, ws.max_row + 1):
            for c in date_cols:
                cell = ws.cell(row=r, column=c)
                if str(cell.value).strip().upper() == str(old_code).strip().upper():
                    cell.value = new_code
                    cell.fill = fill

        wb.save(plan_staff_file)
        return True, "Excel updated with new shift code/color."
    except Exception as e:
        return False, f"Excel update error: {e}"


# ============================================================
# VALIDACIÓN de estructura y salud del archivo (SSoT guardrails)
# ============================================================

def validate_excel_structure(plan_staff_file: str) -> Tuple[bool, List[str], Dict]:
    """
    Valida que el Excel sea una planilla soportada (RGM o Newmont) y que posea columnas de fecha.
    Devuelve: (ok, errors, meta) con meta['variant'] = 'RGM' | 'Newmont' | None y meta['date_columns'].
    """
    errors: List[str] = []
    meta: Dict = {'variant': None, 'date_columns': 0, 'headers': []}

    if not os.path.exists(plan_staff_file):
        errors.append(f"File not found: {plan_staff_file}")
        return False, errors, meta

    try:
        wb = openpyxl.load_workbook(plan_staff_file, read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        errors.append(f"Cannot open workbook: {e}")
        return False, errors, meta

    header_map: Dict[str, int] = {}
    date_count = 0
    for c in ws[1]:
        v = c.value
        if isinstance(v, str):
            header_map[v] = c.column
        elif isinstance(v, datetime):
            date_count += 1

    meta['headers'] = list(header_map.keys())
    meta['date_columns'] = int(date_count)

    # Detectar variante
    variant: Optional[str] = None
    if all(h in header_map for h in ("NAME", "ROLE", "BADGE")):
        variant = "RGM"
    elif all(h in header_map for h in ("Last Name", "First Name", "Discipline", "Company ID")):
        variant = "Newmont"
    else:
        errors.append("Unsupported format. Expected headers for RGM (NAME/ROLE/BADGE) or Newmont (Last Name/First Name/Discipline/Company ID).")

    meta['variant'] = variant  # <- SIEMPRE definido en meta (corrige Pylance)

    if variant is None:
        # Sin variante, no seguimos validando otras reglas
        return False, errors, meta

    if meta['date_columns'] == 0:
        errors.append("No date columns detected in first row (datetime cells).")

    # Confirmar headers mínimos por variante
    if variant == "RGM":
        for h in ("NAME", "ROLE", "BADGE"):
            if h not in header_map:
                errors.append(f"Required header '{h}' missing.")
    else:
        for h in ("Last Name", "First Name", "Discipline", "Company ID"):
            if h not in header_map:
                errors.append(f"Required header '{h}' missing.")

    return (len(errors) == 0), errors, meta


# ============================================================
# Comparación Excel ↔ BD (coherencia con SSoT)
# ============================================================

def check_db_sync_with_excel(plan_staff_file: str, source: str) -> Dict:
    """
    Compara usuarios y schedules entre Excel y BD.
    Retorna:
      {
        'users_in_excel': int,
        'users_in_db': int,
        'missing_badges_in_db': [badge,...],     # En Excel pero NO en BD
        'extra_badges_in_db': [badge,...],       # En BD pero NO en Excel
        'schedule_mismatches': [
          {'badge':..., 'date':'YYYY-MM-DD', 'excel': 'ON', 'db':'OFF'}, ...
        ]
      }
    """
    report = {
        'users_in_excel': 0,
        'users_in_db': 0,
        'missing_badges_in_db': [],
        'extra_badges_in_db': [],
        'schedule_mismatches': []
    }

    ok, _errors, _meta = validate_excel_structure(plan_staff_file)
    if not ok:
        return report

    # --- BD
    try:
        from database_logic import get_all_users, get_schedules_for_source
        users_db = get_all_users(source)
        sched_db = get_schedules_for_source(source)
    except Exception:
        users_db = []
        sched_db = []

    db_badges = {str(u.get('badge', '')).strip() for u in users_db if u.get('badge')}
    report['users_in_db'] = len(db_badges)

    sched_db_map: Dict[str, Dict[str, str]] = {}
    for s in sched_db:
        b = str(s.get('badge', '')).strip()
        d = str(s.get('date', '')).strip()
        st = (s.get('status') or '').strip().upper() if s.get('status') else None
        if not b or not d:
            continue
        sched_db_map.setdefault(b, {})[d] = st

    # --- Excel
    try:
        wb = openpyxl.load_workbook(plan_staff_file, data_only=True)
        ws = wb.active
    except Exception:
        return report

    header_map: Dict[str, int] = {}
    date_cols: Dict[int, date] = {}
    for c in ws[1]:
        v = c.value
        if isinstance(v, str):
            header_map[v] = c.column
        elif isinstance(v, datetime):
            date_cols[c.column] = c.value

    # Determine variant & badge column
    if all(h in header_map for h in ("NAME", "ROLE", "BADGE")):
        badge_col = header_map["BADGE"]
    elif all(h in header_map for h in ("Last Name", "First Name", "Discipline", "Company ID")):
        badge_col = header_map["Company ID"]
    else:
        # No soportado
        return report

    excel_badges: Set[str] = set()
    sched_excel_map: Dict[str, Dict[str, str]] = {}

    def _norm_for_compare(val: object) -> Optional[str]:
        if val is None:
            return None
        s = str(val).strip()
        if s == "":
            return None
        u = s.upper()
        if u in ("OFF", "BREAK", "KO", "LEAVE"):
            return "OFF"
        if u in ("ON", "ON NS"):
            return u
        if u.isdigit() or u == "OK" or "DAY" in u:
            return "ON"
        return u  # personalizado

    for r in range(2, ws.max_row + 1):
        b = ws.cell(row=r, column=badge_col).value
        if not b:
            continue
        b = str(b).strip()
        if not b:
            continue
        excel_badges.add(b)

        # Por fecha
        for c_idx, d in date_cols.items():
            val = ws.cell(row=r, column=c_idx).value
            st = _norm_for_compare(val)
            if st is not None:
                sched_excel_map.setdefault(b, {})[d.isoformat()] = st

    report['users_in_excel'] = len(excel_badges)

    # Diferencias en usuarios
    report['missing_badges_in_db'] = sorted(list(excel_badges - db_badges))
    report['extra_badges_in_db'] = sorted(list(db_badges - excel_badges))

    # Mismatches de schedule (solo comparar cuando ambos tienen valor)
    mismatches: List[Dict] = []
    for b, per_day in sched_excel_map.items():
        for d_str, st_excel in per_day.items():
            st_db = sched_db_map.get(b, {}).get(d_str)
            if st_db is None:
                continue
            if (st_excel or '').upper() != (st_db or '').upper():
                mismatches.append({
                    'badge': b,
                    'date': d_str,
                    'excel': st_excel,
                    'db': st_db
                })
    report['schedule_mismatches'] = mismatches

    return report


# ============================================================
# REGENERACIÓN desde BD (SSoT) — independiente del archivo
# ============================================================

def regenerate_plan_from_db(plan_staff_file: str, source: str) -> Tuple[bool, str]:
    """
    Regenera el archivo PlanStaff (en la ruta indicada) a partir de la BD (SSoT).
    - Si el archivo NO existe o su estructura es inválida, crea una plantilla mínima RGM (NAME/ROLE/BADGE).
    - Luego exporta el estado actual (usuarios+schedules) a dicho archivo.
    """
    try:
        from database_logic import get_all_users, get_schedules_for_source
        users = get_all_users(source)
        schedules = get_schedules_for_source(source)
    except Exception as e:
        return False, f"DB error: {e}"

    # Validar si la plantilla actual es utilizable
    ok, _errors, _meta = validate_excel_structure(plan_staff_file)

    if not ok:
        # Crear plantilla mínima (RGM-like compatible con export_plan_from_db)
        try:
            os.makedirs(os.path.dirname(plan_staff_file), exist_ok=True) if os.path.dirname(plan_staff_file) else None
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Operations_best_opt"
            headers = ["TEAM", "ROLE", "NAME", "BADGE"]
            for col_idx, h in enumerate(headers, start=1):
                ws.cell(row=1, column=col_idx, value=h)
            wb.save(plan_staff_file)
        except Exception as e:
            return False, f"Cannot create template: {e}"

    # Exportar desde BD usando la plantilla (soporta RGM y Newmont si ya existe)
    ok2, msg = export_plan_from_db(plan_staff_file, users, schedules, plan_staff_file, source)
    if ok2:
        return True, f"PlanStaff file regenerated from DB (SSoT): {os.path.basename(plan_staff_file)}"
    else:
        return False, msg


# ============================================================
# REFRESH Excel <- DB (sin sobreescribir celdas ya llenas)
# ============================================================
def refresh_excel_from_db(plan_staff_file: str, source: str) -> Tuple[bool, str]:
    """
    Sincroniza la información desde la BD hacia el Excel existente:
      • Agrega usuarios faltantes (filas) por BADGE.
      • Agrega columnas de fechas que existen en BD y no en el Excel.
      • Escribe solo celdas VACÍAS con el estado proveniente de la BD
        (no modifica valores ya presentes en el Excel).
      • Aplica color y, si el status es un código, comenta 'IN-OUT' (HH:MM-HH:MM).
    """
    ok, _errors, _meta = validate_excel_structure(plan_staff_file)
    if not ok:
        return regenerate_plan_from_db(plan_staff_file, source)

    try:
        from database_logic import get_all_users, get_schedules_for_source, get_shift_type_map
        users_db = get_all_users(source)
        schedules_db = get_schedules_for_source(source)
        custom_map = get_shift_type_map(source)

        wb = openpyxl.load_workbook(plan_staff_file)
        ws = wb.active

        header_map = {cell.value: cell.column for cell in ws[1] if isinstance(cell.value, str)}
        variant = _meta.get('variant')

        date_map: Dict[date, int] = {c.value.date(): c.column for c in ws[1] if isinstance(c.value, datetime)}

        badge_col = header_map.get("BADGE") if variant == "RGM" else header_map.get("Company ID")
        if not badge_col:
            return False, "Badge column not found in Excel."

        # Colores
        green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        yel = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

        def _fill_for(status: Optional[str]) -> Optional[PatternFill]:
            if status is None: return None
            s = str(status).strip().upper()
            if s == "OFF": return red
            if s == "ON NS": return yel
            if s == "ON": return green
            info = custom_map.get(s)
            if info and info.get('color_hex'):
                hex_ = info['color_hex'].lstrip('#').upper()
                return PatternFill(start_color=hex_, end_color=hex_, fill_type="solid")
            return None

        # Build maps for quick lookup
        sched_by_badge: Dict[str, Dict[str, Dict]] = {}
        for s in schedules_db:
            b, d = str(s.get('badge','')).strip(), str(s.get('date','')).strip()
            if b and d:
                sched_by_badge.setdefault(b, {})[d] = s

        rows_by_badge: Dict[str, int] = {str(ws.cell(r,badge_col).value).strip(): r for r in range(2, ws.max_row+1) if ws.cell(r,badge_col).value}

        # --- 1. Add missing users ---
        added_users = 0
        for user in users_db:
            badge = str(user.get('badge','')).strip()
            if badge and badge not in rows_by_badge:
                added_users += 1
                new_row_idx = ws.max_row + 1
                if variant == "RGM":
                    ws.cell(new_row_idx, header_map["NAME"], value=user.get('name',''))
                    ws.cell(new_row_idx, header_map["ROLE"], value=user.get('role',''))
                    ws.cell(new_row_idx, header_map["BADGE"], value=badge)
                else: # Newmont
                    name = user.get('name','')
                    last, first = (name.split(',',1) + [''])[:2] if ',' in name else (name.rsplit(' ',1) + [''])[:2]
                    ws.cell(new_row_idx, header_map["Last Name"], value=last.strip())
                    ws.cell(new_row_idx, header_map["First Name"], value=first.strip())
                    ws.cell(new_row_idx, header_map["Discipline"], value=user.get('role',''))
                    ws.cell(new_row_idx, header_map["Company ID"], value=badge)
                rows_by_badge[badge] = new_row_idx

        # --- 2. Add missing date columns ---
        all_db_dates = {datetime.fromisoformat(d).date() for b in sched_by_badge for d in sched_by_badge[b]}
        missing_dates = sorted(list(all_db_dates - set(date_map.keys())))
        for d in missing_dates:
            new_col = ws.max_column + 1
            ws.cell(1, new_col, value=datetime.combine(d, datetime.min.time()))
            date_map[d] = new_col

        # --- 3. Fill empty cells ---
        filled_cells = 0
        for badge, row_idx in rows_by_badge.items():
            user_scheds = sched_by_badge.get(badge, {})
            for d, col_idx in date_map.items():
                cell = ws.cell(row_idx, col_idx)
                if cell.value is None or str(cell.value).strip() == '':
                    db_info = user_scheds.get(d.isoformat())
                    if db_info:
                        status = (db_info.get('status') or '').strip().upper()
                        if status:
                            filled_cells += 1
                            cell.value = status
                            cell.fill = _fill_for(status) or PatternFill(fill_type=None)
                            if status not in ("ON","ON NS","OFF") and db_info.get('in_time') and db_info.get('out_time'):
                                cell.comment = Comment(f"{db_info['in_time']}-{db_info['out_time']}", "ShiftType")

        wb.save(plan_staff_file)
        return True, f"Refresh complete: added {added_users} users and filled {filled_cells} cells from DB."

    except Exception as e:
        return False, f"Refresh error: {e}"